import cv2
import os
from flask import Flask, request, render_template
from datetime import date, datetime
import numpy as np
from sklearn.neighbors import KNeighborsClassifier
import pandas as pd
import joblib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)

nimgs = 100
imgBackground = cv2.imread("background.png")

datetoday = date.today().strftime("%m_%d_%y")
datetoday2 = date.today().strftime("%d-%B-%Y")

face_detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')

# Ensure required directories exist
os.makedirs('Attendance', exist_ok=True)
os.makedirs('static/faces', exist_ok=True)

if f'Attendance-{datetoday}.csv' not in os.listdir('Attendance'):
    with open(f'Attendance/Attendance-{datetoday}.csv', 'w') as f:
        f.write('Name,Roll,Time\n')


def totalreg():
    return len(os.listdir('static/faces'))


def extract_faces(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return face_detector.detectMultiScale(gray, 1.2, 5, minSize=(20, 20))


def identify_face(facearray):
    model = joblib.load('static/face_recognition_model.pkl')
    return model.predict(facearray)


def train_model():
    faces, labels = [], []
    userlist = os.listdir('static/faces')
    for user in userlist:
        for imgname in os.listdir(os.path.join('static/faces', user)):
            img_path = os.path.join('static/faces', user, imgname)
            img = cv2.imread(img_path)
            if img is not None:
                resized_face = cv2.resize(img, (50, 50))
                faces.append(resized_face.ravel())
                labels.append(user)
    if faces:
        knn = KNeighborsClassifier(n_neighbors=5)
        knn.fit(np.array(faces), labels)
        joblib.dump(knn, 'static/face_recognition_model.pkl')


def extract_attendance():
    try:
        df = pd.read_csv(f'Attendance/Attendance-{datetoday}.csv')
        names, rolls, times = df['Name'], df['Roll'], df['Time']
        return names, rolls, times, len(df)
    except:
        return [], [], [], 0


def add_attendance(name):
    username, userid = name.split('_')
    current_time = datetime.now().strftime("%H:%M:%S")

    csv_path = f'Attendance/Attendance-{datetoday}.csv'
    excel_path = f'Attendance/Attendance-{datetoday}.xlsx'

    # --- Step 1: Ensure CSV file exists ---
    if not os.path.exists(csv_path):
        df = pd.DataFrame(columns=['Name', 'Roll', 'Time'])
        df.to_csv(csv_path, index=False)
    else:
        df = pd.read_csv(csv_path)

    # --- Step 2: Prevent duplicate entries ---
    if int(userid) in df['Roll'].values:
        print(f"[INFO] {username} already marked present today.")
        return  # Skip duplicate

    # --- Step 3: Add new entry ---
    new_entry = pd.DataFrame([[username, int(userid), current_time]],
                             columns=['Name', 'Roll', 'Time'])
    df = pd.concat([df, new_entry], ignore_index=True)

    # --- Step 4: Save updated CSV ---
    df.to_csv(csv_path, index=False)
    print(f"[INFO] Attendance marked for {username} at {current_time}.")

    # --- Step 5: Save to Excel with formatting ---
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Add header
        headers = ["Name", "Roll", "Time"]
        ws.append(headers)

        # Header style
        for col in ws[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.alignment = Alignment(horizontal="center")
            col.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Add rows
        for _, row in df.iterrows():
            ws.append(list(row))

        # Auto-adjust column width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 3

        # Save Excel
        wb.save(excel_path)
        print(f"[INFO] Excel updated successfully → {excel_path}")

    except Exception as e:
        print(f"[ERROR] Failed to update Excel file: {e}")

@app.route('/')
def home():
    names, rolls, times, l = extract_attendance()
    return render_template('home.html', names=names, rolls=rolls, times=times, l=l,
                           totalreg=totalreg(), datetoday2=datetoday2)


@app.route('/start')
def start():
    names, rolls, times, l = extract_attendance()

    model_path = 'static/face_recognition_model.pkl'
    if not os.path.exists(model_path):
        return render_template('home.html', names=names, rolls=rolls, times=times, l=l,
                               totalreg=totalreg(), datetoday2=datetoday2,
                               mess='⚠️ No trained model found. Please add a new user first.')

    model = joblib.load(model_path)
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        return render_template('home.html', names=names, rolls=rolls, times=times, l=l,
                               totalreg=totalreg(), datetoday2=datetoday2,
                               mess='⚠️ Camera not accessible.')

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        faces = extract_faces(frame)
        for (x, y, w, h) in faces:
            face = cv2.resize(frame[y:y+h, x:x+w], (50, 50)).reshape(1, -1)
            identified_person = model.predict(face)[0]
            add_attendance(identified_person)

            cv2.rectangle(frame, (x, y), (x+w, y+h), (50, 50, 255), 2)
            cv2.putText(frame, identified_person, (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

        imgBackground[162:162 + 480, 55:55 + 640] = frame
        cv2.imshow('Attendance', imgBackground)
        if cv2.waitKey(1) == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

    names, rolls, times, l = extract_attendance()
    return render_template('home.html', names=names, rolls=rolls, times=times, l=l,
                           totalreg=totalreg(), datetoday2=datetoday2)


@app.route('/add', methods=['POST'])
def add():
    newusername = request.form['newusername']
    newuserid = request.form['newuserid']
    user_dir = os.path.join('static/faces', f'{newusername}_{newuserid}')
    os.makedirs(user_dir, exist_ok=True)

    cap = cv2.VideoCapture(0)
    count = 0
    while count < nimgs:
        ret, frame = cap.read()
        faces = extract_faces(frame)
        for (x, y, w, h) in faces:
            face_img = frame[y:y+h, x:x+w]
            filename = os.path.join(user_dir, f'{newusername}_{count}.jpg')
            cv2.imwrite(filename, face_img)
            count += 1
            cv2.putText(frame, f'Captured: {count}/{nimgs}', (20, 40),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
        cv2.imshow('Registering User', frame)
        if cv2.waitKey(1) == 27:
            break

    cap.release()
    cv2.destroyAllWindows()
    train_model()

    names, rolls, times, l = extract_attendance()
    return render_template('home.html', names=names, rolls=rolls, times=times, l=l,
                           totalreg=totalreg(), datetoday2=datetoday2,
                           mess=f'✅ User {newusername} added & model retrained.')


if __name__ == '__main__':
    app.run(debug=True)
